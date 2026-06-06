# import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

""" 配置excel 表头颜色，宽度，高度，等信息 """
field_config = {
    "Date": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Time": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Amp.Temp": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "E_out": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Placeholder": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "SOC/%": {"width": 13,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "BESS Auxiliary Power [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Common Infrastructure Aux Power[MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "PV Power To Plant Substation BCP [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "BESS Power To Plant Substation BCP [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Mode": {"width": 12,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "PV Power to BESS Plant BCP  [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "DisCharge Power BCP [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "PV Power To BESS plant aux consumuer [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "PV Power To Common Infrastructure Power [MW] ": {"width": 25,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Charge Power [DC]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Discharge Power [DC]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "BESS Power To BESS plant aux consumuer [MW]": {"width": 25,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "BESS plant Power To Common Infrastructure Power [MW]": {"width": 25,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Imported Power [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "BESS Power To PV Plant [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Energy [DC]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
    "Exported Power [MW]": {"width": 22,  "fill_color": None, "header_fill_color": "BDD7EE"},
}

# 对表格进行处理，
def save_excel_with_field_config_pv(
    save_path,
    sheet_name="Sheet1", # 表的名称
    field_config=field_config,
    default_width=15, # 表头的高度，统一值
    header_fill_color="BDD7EE",      # 全局表头背景色（当字段未单独指定时使用）
    header_font_color="000000",      # 全局字体的颜色
    header_height=50,
    wrap_header=True,
    wrap_data=True,
    debug=False,
    add_GWh=False,
    add_GWh_col=1,
):
    """
    将已存在的 Excel 文件按 field_config 格式化：
    - 表头行：支持全局或每个字段单独设置背景色、字体、边框、高度、自动换行
    - 数据行：每个字段可单独设置背景色、自动换行
    - 列宽：每个字段可单独设置宽度
    """
    if field_config is None:
        field_config = {}

    wb = load_workbook(save_path)
    ws = wb[sheet_name]

    # ----- 1. 获取表头映射：字段名 -> 列索引（去除首尾空格，便于匹配）-----
    headers_raw = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            # 存储原始值（去除空格后的字符串）作为 key
            key = str(cell.value).strip()
            headers_raw[key] = col_idx

    # ----- 2. 设置表头样式（支持按字段单独配置背景色）-----
    # 全局表头样式（字体、对齐、边框等）
    header_font = Font(color=header_font_color, bold=True)
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=wrap_header
    )
    black_side = Side(style="thin", color="000000")
    header_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)


    # 遍历第一行每个单元格，应用样式
    for col_idx, cell in enumerate(ws[1], start=1):
        # 获取当前字段名（去除空格）
        field_name_raw = cell.value
        if field_name_raw is None:
            # 如果表头单元格为空，使用全局背景色
            fill_color = header_fill_color
        else:
            field_name = str(field_name_raw).strip()
            cfg = field_config.get(field_name, {})
            # 优先使用字段单独指定的表头背景色，否则使用全局
            fill_color = cfg.get("header_fill_color", header_fill_color)

        # 应用样式
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    ws.row_dimensions[1].height = header_height

    # ----- 3. 设置列宽 -----
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        field_name_raw = ws.cell(row=1, column=col_idx).value
        if field_name_raw is None:
            width = default_width
        else:
            field_name = str(field_name_raw).strip()
            cfg = field_config.get(field_name, {})
            width = cfg.get("width", default_width)
        ws.column_dimensions[col_letter].width = width

    # ----- 4. 设置数据区域样式（背景色 + 自动换行）-----
    max_row = ws.max_row
    if wrap_data:
        data_alignment = Alignment(wrap_text=True)

    # 预先建立字段名到列索引的映射（去除空格后）
    for field_name, cfg in field_config.items():
        # 去除 field_name 首尾空格，与 headers_raw 中的 key 一致
        clean_field = field_name.strip()
        col_idx = headers_raw.get(clean_field)
        if col_idx is None:
            if debug:
                print(f"未找到字段: {field_name}")
            continue

        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # 数据区域背景色（使用 fill_color）
        fill_color = cfg.get("fill_color")
        fill = PatternFill(fill_type="solid", fgColor=fill_color) if fill_color else None

        for row in range(2, max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if fill:
                cell.fill = fill
            if wrap_data:
                cell.alignment = data_alignment
    # 对于拼接的地方，进行颜色填充
    if add_GWh is True:
        v1 = int(add_GWh_col)
        v2 = int(v1 + 1)
        v3 = int(v1 + 2)

        config = {
            v1: {'rows': [2], 'fill_color': '4A86E8', 'width': 15},
            v2: {'rows': [2, 3, 4,5], 'fill_color': '4A86E8', 'width': 30},
            v3: {'rows': [2, 3, 4,5], 'fill_color': '4A86E8', 'width': 15}
        }
        for col_idx, cfg in config.items():
            fill_color = cfg.get('fill_color', None)
            width = cfg.get('width', None)
            rows = cfg.get('rows', [])

            if width:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = width

            if fill_color:
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                for row in rows:
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(save_path)

# 对统计的数值gwh 表格进行颜色填充
def save_excel_with_field_config_GWh(file_path, sheet_name='Sheet2',
                                     row_height=15,
                                     col_widths={'A': 10, 'B': 31, 'C': 15},
                                     fill_color='4A86E8'):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 设置行高
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = row_height

    # 设置列宽
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    blue_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # 遍历所有单元格
    for row in ws.iter_rows():
        for cell in row:
            # 设置居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 判断是否填充：A1 或者 单元格值为数字
            if (cell.row == 1 and cell.column == 1) or isinstance(cell.value, (int, float)) or isinstance(cell.value, str):
                cell.fill = blue_fill
            # 否则不填充（保持原样，不清除）
    wb.save(file_path)



from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PHASE_NAMES = ["Phase A", "Phase B", "Phase C"]
AUX_HEADERS = ["Temperature", "Aux power operation MW", "Aux power idle MW"]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def _json_value(value: Any, field_name: str | None = None) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        numeric_text = value.replace(",", "").replace("%", "").strip()
        try:
            numeric_value = float(numeric_text)
        except ValueError:
            return value
    else:
        numeric_value = value

    if isinstance(numeric_value, (int, float)):
        result = float(numeric_value)
        if field_name and "%" in field_name and result > 1:
            result = result / 100
        return int(result) if result.is_integer() else result

    return value


def _temperature_value(value: Any) -> Any:
    text = _clean_text(value).replace("℃", "").replace("°C", "").strip()
    if not text:
        return None
    try:
        result = float(text)
    except ValueError:
        return text
    return int(result) if result.is_integer() else result


def _find_design_workbook(root: Path) -> Path:
    data_dir = next(p for p in root.iterdir() if p.is_dir() and p.name.startswith("PVsyst"))

    candidates = []
    for path in data_dir.rglob("*.xlsx"):
        name = _clean_text(path.name)
        if name.startswith("bak_") or name.startswith("~$"):
            continue
        if name.startswith("DEWA7") and "BESS_PPM" in name and "06-04-2026" in name:
            candidates.append(path)

    if not candidates:
        raise FileNotFoundError("Cannot find DEWA7 BESS PPM design sheet.")
    if len(candidates) > 1:
        candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _read_phase_sheet(ws) -> dict[str, dict[str, dict[str, Any]]]:
    result: dict[str, dict[str, dict[str, Any]]] = {}

    for row in ws.iter_rows():
        for cell in row:
            phase_name = _clean_text(cell.value)
            if phase_name not in PHASE_NAMES:
                continue

            phase_col = cell.column
            year_col = phase_col
            capacity_col = phase_col + 1
            battery_efficiency_col = phase_col + 2

            rows = {}
            for row_idx in range(cell.row + 1, ws.max_row + 1):
                year = _clean_text(ws.cell(row_idx, year_col).value)
                capacity = ws.cell(row_idx, capacity_col).value
                battery_efficiency = ws.cell(row_idx, battery_efficiency_col).value

                if not year:
                    continue
                if year in {"Year", "Middle of Year"} or "Capacity" in year:
                    continue

                capacity_value = _json_value(capacity)
                battery_efficiency_value = _json_value(battery_efficiency, "Battery Efficiency [%]:")
                rows[year] = {
                    "Functioanl Capacity": capacity_value,
                    "capacity": capacity_value,
                    "Battery Efficiency [%]:": battery_efficiency_value,
                }

            result[phase_name] = rows

    return result


def _add_eta_values(params: dict[str, Any]) -> None:
    for phase_name, year_values in params["Phase"].items():
        ppm = params["PPM Design Input Sheet"].get(phase_name, {})
        fixed_efficiencies = [
            ppm.get("DC Cable Efficiency [%]:"),
            ppm.get("PCS Efficiency [%]:"),
            ppm.get("LV Cable Efficiency [%]:"),
            ppm.get("MV Transformer Efficiency [%]:"),
            ppm.get("MV Cable Efficiency [%]:"),
        ]

        if any(value is None for value in fixed_efficiencies):
            continue

        eta_dis = 1
        for value in fixed_efficiencies:
            eta_dis *= value

        for row in year_values.values():
            row["eta_dis"] = eta_dis
            battery_efficiency = row.get("Battery Efficiency [%]:")
            if battery_efficiency is None:
                row["eta_ch"] = None
            else:
                row["eta_ch"] = eta_dis * battery_efficiency


def _is_stage_name(value: Any) -> bool:
    text = _clean_text(value)
    return text == "CY-01" or re.fullmatch(r"Stage\s*\d+", text) is not None


def _normalize_stage_name(value: Any) -> str:
    text = _clean_text(value)
    match = re.fullmatch(r"Stage\s*(\d+)", text)
    if match:
        return f"Stage {match.group(1)}"
    return text


def _find_aux_title(ws, start_row: int, phase_name: str) -> tuple[int, int] | None:
    title_prefix = f"{phase_name} BESS Aux Power"
    for row_idx in range(start_row, min(start_row + 8, ws.max_row) + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = _clean_text(ws.cell(row_idx, col_idx).value)
            if value.startswith(title_prefix):
                return row_idx, col_idx
    return None


def _read_aux_table(ws) -> dict[str, dict[str, dict[str, list[Any]]]]:
    result: dict[str, dict[str, dict[str, list[Any]]]] = {}
    seen_stage_rows: set[int] = set()

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value
            if not _is_stage_name(value):
                continue

            stage_name = _normalize_stage_name(value)
            if row_idx in seen_stage_rows:
                continue

            phase_titles = {
                phase_name: _find_aux_title(ws, row_idx, phase_name)
                for phase_name in PHASE_NAMES
            }
            if not any(phase_titles.values()):
                continue

            seen_stage_rows.add(row_idx)
            result[stage_name] = {}

            for phase_name, position in phase_titles.items():
                if position is None:
                    continue
                title_row, temp_col = position
                header_row = title_row + 1
                data_start = header_row + 1
                data = {header: [] for header in AUX_HEADERS}

                for data_row in range(data_start, ws.max_row + 1):
                    temp = _temperature_value(ws.cell(data_row, temp_col).value)
                    if temp is None:
                        break
                    data["Temperature"].append(temp)
                    data["Aux power operation MW"].append(
                        _json_value(ws.cell(data_row, temp_col + 1).value)
                    )
                    data["Aux power idle MW"].append(
                        _json_value(ws.cell(data_row, temp_col + 2).value)
                    )

                result[stage_name][phase_name] = data

    return result


def _read_ppm_sheet(ws) -> dict[str, dict[str, Any]]:
    result = {phase_name: {} for phase_name in PHASE_NAMES}

    for row_idx in range(1, ws.max_row + 1):
        field_name = _clean_text(ws.cell(row_idx, 1).value)
        if not field_name:
            continue

        phase_values = [ws.cell(row_idx, col_idx).value for col_idx in (2, 3, 4)]
        normalized = [_clean_text(v) for v in phase_values]

        if normalized == PHASE_NAMES:
            continue
        if all(v is None or _clean_text(v) == "" for v in phase_values):
            continue

        for phase_name, value in zip(PHASE_NAMES, phase_values):
            parsed = _json_value(value, field_name)
            if parsed is not None:
                result[phase_name][field_name] = parsed

    return result


def read_design_sheet(workbook_path: Path | None = None) -> dict[str, Any]:
    root = Path(__file__).resolve().parents[1]
    workbook_path = workbook_path or _find_design_workbook(root)

    wb = load_workbook(workbook_path, data_only=True)
    params = {
        "source_file": str(workbook_path),
        "Phase": _read_phase_sheet(wb["Phase"]),
        "AuxPower6h": _read_aux_table(wb["AuxPower6h"]),
        "Aux power 8h": _read_aux_table(wb["Aux power 8h"]),
        "PPM Design Input Sheet": _read_ppm_sheet(wb["PPM Design Input Sheet"]),
    }
    _add_eta_values(params)
    wb.close()
    return params


def save_design_sheet_params(params: dict[str, Any], output_path: Path | None = None) -> Path:
    root = Path(__file__).resolve().parents[1]
    output_path = output_path or root / "datavalues.json"
    output_path.write_text(
        json.dumps(params, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


def get_design_sheet_params(
        workbook_path: Path | str | None = None,
        output_path: Path | str | None = None,
        save: bool = True,
) -> dict[str, Any]:
    params = read_design_sheet(Path(workbook_path) if workbook_path else None)
    if save:
        save_design_sheet_params(
            params,
            Path(output_path) if output_path else None,
        )
    return params


# def main() -> dict[str, Any]:
#     root = Path(__file__).resolve().parents[1]
#     output_path = root / "datavalues.json"
#     params = get_design_sheet_params(output_path=output_path, save=True)
#     print(f"Saved {output_path}")
#     return params
#
#
# if __name__ == "__main__":
#     main()

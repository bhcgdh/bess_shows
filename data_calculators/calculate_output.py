from openpyxl import load_workbook, Workbook
import os
import pandas as pd
import numpy as np
import time
import xlsxwriter
from data_outputs.outputs_excel_config_type2 import format_excel_all_sheets
from data_outputs.output_pv_save import _deal_save_data_df_pv_sheet1

#  使用表格是循环生成的27张表， R"F:\bess_shows\PVsyst数据\结果数据\Phase_A-27 years_260406"
# 修改输出，对应表格参考 R"F:\bess_shows\PVsyst数据\UAE_DEWA7_PhaseA_PM_Output v1.0.xlsx"
# 设置 表格的字段名称， 双轴
# 修改 20260528 - 更新计算方式，和两个列的命名方式   AC MV Aux Power > AC LV Aux Power
def cal_ouput_columns():
    columns_multi = [
        ('Time', 'Time Step'),
        ('Time', 'Hour'),
        ('Time', 'Day'),
        ('Time', 'Month'),
        ('PV', 'AC MV Power Available'),
        ('PV', 'AC MV Surplus'),
        ('PV', 'AC MV Power to Infra'),
        ('PV', 'AC MV Power to BESS'),
        ('PV', 'AC MV Power to HV'),
        ('PV', 'AC MV Power Dumped'),
        ('BESS', 'AC MV Power Charge'),
        ('BESS', 'AC MV Power Discharge'),
        ('BESS', 'DC Power Charge'),
        ('BESS', 'DC Power Discharge'),
        ('BESS', 'DC Energy'),
        ('BESS', 'SOC'),
        ('BESS', 'Mode'),
        ('BESS', 'AC LV Aux Power'),
        ('BESS', 'AC MV Power to Infra'),
        ('BESS', 'AC LV Power to BESS'),
        ('BESS', 'AC MV Power to PV'),
        ('BESS', 'AC MV Power to HV'),
        ('HV', 'AC HV Power at MV of MV/HV Transformer'),
        ('HV', 'AC HV Power at EDP'),
        ('HV', 'Exported Power at EDP'),
        ('HV', 'Imported Power at EDP'),
        ('HVs', 'Exported Power at EDP.1'),
        ('HVs', 'Exported Power > 1,000 MW'),
        ('HVs','Exported Power from BESS at EDP'),
        ('HVs', 'Exported Power from PV at EDP'),
    ]

    # 创建 MultiIndex
    col = pd.MultiIndex.from_tuples(columns_multi)
    return col

# 比较少用 ，
# 表格中第三行，数据计算
def _cal_output_cell3(df, result=None):
    """
    计算 cell3 各类 PV / BESS / HV 指标，字段使用 MultiIndex 格式
    """
    if result is None:
        result = {}
    result['cell3'] = {}

    # Time 类
    result['cell3']['Time'] = {}
    result['cell3']['Time']["Time Step"] = "[-]"
    result['cell3']['Time']["Hour"] = "[-]"
    result['cell3']['Time']["Day"] = "[-]"
    result['cell3']['Time']["Month"] = "[-]"


    # PV 类
    result['cell3']['PV'] = {}
    result['cell3']['PV'][('PV', 'AC MV Power Available')] = df[df[('PV', 'AC MV Power Available')] > 0][('PV', 'AC MV Power Available')].sum()
    result['cell3']['PV'][('PV', 'AC MV Surplus')] = df[('PV', 'AC MV Surplus')].min()
    result['cell3']['PV'][('PV', 'AC MV Power to Infra')] = df[('PV', 'AC MV Power to Infra')].sum()
    result['cell3']['PV'][('PV', 'AC MV Power to BESS')] = df[('PV', 'AC MV Power to BESS')].min()
    result['cell3']['PV'][('PV', 'AC MV Power to HV')] = df[('PV', 'AC MV Power to HV')].min()
    result['cell3']['PV'][('PV', 'AC MV Power Dumped')] = df[('PV', 'AC MV Power Dumped')].min()

    # BESS 类
    result['cell3']['BESS'] = {}
    result['cell3']['BESS'][('BESS', 'AC MV Power Charge')] = df[('BESS', 'AC MV Power Charge')].min()
    result['cell3']['BESS'][('BESS', 'AC MV Power Discharge')] = df[('BESS', 'AC MV Power Discharge')].min()
    result['cell3']['BESS'][('BESS', 'DC Power Charge')] = df[('BESS', 'DC Power Charge')].min()
    result['cell3']['BESS'][('BESS', 'DC Power Discharge')] = df[('BESS', 'DC Power Discharge')].min()
    result['cell3']['BESS'][('BESS', 'DC Energy')] = df[('BESS', 'DC Energy')].min()

    tmp = [float(i.split("%")[0]) for i in list(df[('BESS', 'SOC')])]
    result['cell3']['BESS'][('BESS', 'SOC')] = np.min(tmp)

    result['cell3']['BESS'][('BESS', 'Mode')] = df[df[('BESS', 'Mode')] == '1'].shape[0]

    result['cell3']['BESS'][('BESS', 'AC LV Aux Power')] = df[('BESS', 'AC LV Aux Power')].sum()
    result['cell3']['BESS'][('BESS', 'AC MV Power to Infra')] = df[('BESS', 'AC MV Power to Infra')].sum()
    result['cell3']['BESS'][('BESS', 'AC LV Power to BESS')] = df[('BESS', 'AC LV Power to BESS')].sum()
    result['cell3']['BESS'][('BESS', 'AC MV Power to PV')] = df[('BESS', 'AC MV Power to PV')].sum()
    result['cell3']['BESS'][('BESS', 'AC MV Power to HV')] = df[('BESS', 'AC MV Power to HV')].min()

    # HV 类
    result['cell3']['HV'] = {}
    result['cell3']['HV'][('HV', 'AC HV Power at MV of MV/HV Transformer')] = df[('HV', 'AC HV Power at MV of MV/HV Transformer')].min()
    result['cell3']['HV'][('HV', 'AC HV Power at EDP')] = df[('HV', 'AC HV Power at EDP')].min()
    result['cell3']['HV'][('HV', 'Exported Power at EDP')] = df[('HV', 'Exported Power at EDP')].min()
    result['cell3']['HV'][('HV', 'Imported Power at EDP')] = df[('HV', 'Imported Power at EDP')].min()

    # HVs 类
    result['cell3']['HVs'] = {}
    result['cell3']['HVs'][('HVs', 'Exported Power at EDP.1')] = df[('HVs', 'Exported Power at EDP.1')].min()
    result['cell3']['HVs'][('HVs', 'Exported Power > 1,000 MW')] = df[('HVs', 'Exported Power > 1,000 MW')].min()
    result['cell3']['HVs'][('HVs', 'Exported Power from BESS at EDP')] = df[('HVs', 'Exported Power from BESS at EDP')].sum()
    result['cell3']['HVs'][('HVs', 'Exported Power from PV at EDP')] = df[('HVs', 'Exported Power from PV at EDP')].sum()

    return result

# 表格中第四行，数据计算
def _cal_output_cell4(df, result=None):
    """
    计算 cell4 各类 PV / BESS / HV 指标（最大值 / 百分比）
    """
    if result is None:
        result = {}
    result['cell4'] = {}

    # Time 类
    result['cell4']['Time'] = {}
    result['cell4']['Time']["Time Step"] = "[-]"
    result['cell4']['Time']["Hour"] = "[-]"
    result['cell4']['Time']["Day"] = "[-]"
    result['cell4']['Time']["Month"] = "[-]"

    # PV 类
    result['cell4']['PV'] = {}
    result['cell4']['PV'][('PV', 'AC MV Power Available')] = df[('PV', 'AC MV Power Available')].max()
    result['cell4']['PV'][('PV', 'AC MV Surplus')] = df[('PV', 'AC MV Surplus')].max()
    result['cell4']['PV'][('PV', 'AC MV Power to Infra')] = df[('PV', 'AC MV Power to Infra')].max()
    result['cell4']['PV'][('PV', 'AC MV Power to BESS')] = df[('PV', 'AC MV Power to BESS')].max()
    result['cell4']['PV'][('PV', 'AC MV Power to HV')] = df[('PV', 'AC MV Power to HV')].max()
    result['cell4']['PV'][('PV', 'AC MV Power Dumped')] = df[('PV', 'AC MV Power Dumped')].max()

    # BESS 类
    result['cell4']['BESS'] = {}
    result['cell4']['BESS'][('BESS', 'AC MV Power Charge')] = df[('BESS', 'AC MV Power Charge')].max()
    result['cell4']['BESS'][('BESS', 'AC MV Power Discharge')] = df[('BESS', 'AC MV Power Discharge')].max()
    result['cell4']['BESS'][('BESS', 'DC Power Charge')] = df[('BESS', 'DC Power Charge')].max()
    result['cell4']['BESS'][('BESS', 'DC Power Discharge')] = df[('BESS', 'DC Power Discharge')].max()
    result['cell4']['BESS'][('BESS', 'DC Energy')] = df[('BESS', 'DC Energy')].max()

    tmp = [float(i.split("%")[0]) for i in list(df[('BESS', 'SOC')])]
    result['cell4']['BESS'][('BESS', 'SOC')] = np.max(tmp)

    result['cell4']['BESS'][('BESS', 'Mode')] = df[df[('BESS', 'Mode')] == '1'].shape[0] / 365 #问题，原来是 1
    result['cell4']['BESS'][('BESS', 'AC LV Aux Power')] = df[('BESS', 'AC LV Aux Power')].max()
    result['cell4']['BESS'][('BESS', 'AC MV Power to Infra')] = df[('BESS', 'AC MV Power to Infra')].max()
    result['cell4']['BESS'][('BESS', 'AC LV Power to BESS')] = df[('BESS', 'AC LV Power to BESS')].max()
    result['cell4']['BESS'][('BESS', 'AC MV Power to PV')] = df[('BESS', 'AC MV Power to PV')].max()
    result['cell4']['BESS'][('BESS', 'AC MV Power to HV')] = df[('BESS', 'AC MV Power to HV')].max()

    # HV 类
    result['cell4']['HV'] = {}
    result['cell4']['HV'][('HV', 'AC HV Power at MV of MV/HV Transformer')] = df[('HV', 'AC HV Power at MV of MV/HV Transformer')].max()
    result['cell4']['HV'][('HV', 'AC HV Power at EDP')] = df[('HV', 'AC HV Power at EDP')].max()
    result['cell4']['HV'][('HV', 'Exported Power at EDP')] = df[('HV', 'Exported Power at EDP')].max()
    result['cell4']['HV'][('HV', 'Imported Power at EDP')] = df[('HV', 'Imported Power at EDP')].max()

    # HVs 类
    result['cell4']['HVs'] = {}
    result['cell4']['HVs'][('HVs', 'Exported Power at EDP.1')] = df[('HVs', 'Exported Power at EDP.1')].max()
    result['cell4']['HVs'][('HVs', 'Exported Power > 1,000 MW')] = df[('HVs', 'Exported Power > 1,000 MW')].max()
    result['cell4']['HVs'][('HVs', 'Exported Power from BESS at EDP')] = df[('HVs', 'Exported Power from BESS at EDP')].max()
    result['cell4']['HVs'][('HVs', 'Exported Power from PV at EDP')] = df[('HVs', 'Exported Power from PV at EDP')].max()

    return result

# 表格中第五行，数据计算
def _cal_output_cell5(result=None):
    # 初始化 cell5
    if result==None:
        result={}
    result['cell5'] = {}

    # Time 类
    result['cell5']['Time'] = {}
    result['cell5']['Time']["Time Step"] = "[-]"
    result['cell5']['Time']["Hour"] = "[-]"
    result['cell5']['Time']["Day"] = "[-]"
    result['cell5']['Time']["Month"] = "[-]"

    # PV 类
    result['cell5']['PV'] = {}
    result['cell5']['PV']["AC MV Power Available"] = "[MW]"
    result['cell5']['PV']["AC MV Surplus"] = "[MW]"
    result['cell5']['PV']["AC MV Power to Infra"] = "[MW]"
    result['cell5']['PV']["AC LV Power to BESS"] = "[MW]"
    result['cell5']['PV']["AC MV Power to PV"] = "[MW]"
    result['cell5']['PV']["AC MV Power to HV"] = "[MW]"
    result['cell5']['PV']["AC MV Power Dumped"] = "[MW]"

    # BESS 类
    result['cell5']['BESS'] = {}
    result['cell5']['BESS']["AC MV Power Charge"] = "[MW]"
    result['cell5']['BESS']["AC MV Power Discharge"] = "[MW]"
    result['cell5']['BESS']["DC Power Charge"] = "[MW]"
    result['cell5']['BESS']["DC Power Discharge"] = "[MW]"
    result['cell5']['BESS']["DC Energy"] = "[MWh]"
    result['cell5']['BESS']["SOC"] = "[%]"
    result['cell5']['BESS']["Mode"] = "[-]"
    result['cell5']['BESS']["AC LV Aux Power"] = "[MW]"
    result['cell5']['BESS']["AC MV Power to Infra"] = "[MW]"
    result['cell5']['BESS']["AC LV Power to BESS"] = "[MW]"
    result['cell5']['BESS']["AC MV Power to PV"] = "[MW]"
    result['cell5']['BESS']["AC MV Power to HV"] = "[MW]"

    # HV 类
    result['cell5']['HV'] = {}
    result['cell5']['HV']['AC HV Power at MV of MV/HV Transformer'] = "[MW]"
    result['cell5']['HV']["AC HV Power at EDP"] = "[MW]"
    result['cell5']['HV']["Exported Power at EDP"] = "[MW]"
    result['cell5']['HV']["Imported Power at EDP"] = "[MW]"

    result['cell5']['HVs'] = {}
    result['cell5']['HVs']["Exported Power at EDP.1"] = "[MW]"
    result['cell5']['HVs']["Exported Power > 1,000 MW"] = "[MW]"
    result['cell5']['HVs']['Exported Power from BESS at EDP'] = "[MW]"
    result['cell5']['HVs'][ 'Exported Power from PV at EDP'] = "[MW]"


    return result
def safe_concat(dfs, ignore_index=True):
    # 排除空 DataFrame 或 所有元素都是 NA 的 DataFrame
    valid_dfs = [
        df for df in dfs
        if not (df.empty or df.isna().all().all())
    ]
    if not valid_dfs:
        return pd.DataFrame()
    return pd.concat(valid_dfs, ignore_index=ignore_index)

def _build_cell_row(df, result_cell):
    new_row = []
    for col in df.columns:
        val = None
        if col[0] in result_cell:
            inner_dict = result_cell[col[0]]
            val = inner_dict.get(col)
            if val is None:
                val = inner_dict.get(col[1])
        new_row.append(val)
    return new_row

#  数据插入指定的行
def _insert_cell_to_df(df, result_cell5, row_index=3):
    """
    在 DataFrame 的指定行索引处插入新行（不覆盖现有行），新行内容来自 result_cell5。

    参数:
        df: pandas DataFrame，列是 MultiIndex [(col1, col2), ...]
        result_cell5: dict，格式如 result['cell5']
        row_index: int，要插入的行索引（从0开始）。现有行及以下行将向下移动。

    返回:
        df: 更新后的 DataFrame
    """
    # 不用全列转 object，直接插入字符串/数值都可
    new_row = []

    for col in df.columns:  # col 是 tuple (lvl0, lvl1)
        val = None
        if col[0] in result_cell5:  # 一级 key
            inner_dict = result_cell5[col[0]]
            # 宽松匹配 tuple 或 str
            for k, v in inner_dict.items():
                if k == col:  # tuple 完全匹配
                    val = v
                    break
                # 可兼容字符串 key（cell5 的情况）
                if isinstance(k, str) and str(k).strip() == str(col[1]).strip():
                    val = v
                    break
        new_row.append(val)

    new_row_df = pd.DataFrame([new_row], columns=df.columns)
    # 会导致
    # new_row_df = pd.DataFrame(
    #     [[""] * len(df.columns)],
    #     columns=df.columns
    # )
    if row_index >= len(df):
        df = pd.concat([df, new_row_df], ignore_index=True)
        # df = safe_concat([df, new_row_df])
    else:
        df1 = df.iloc[:row_index, :]
        df2 = df.iloc[row_index:, :]
        # df = safe_concat([df1, new_row_df, df2])
        df = pd.concat([df1, new_row_df, df2], ignore_index=True)
    return df

#  标投增加数据的计算统计
def cal_output_head(df):
    result3 = _cal_output_cell3(df)
    result4 = _cal_output_cell4(df)
    result5 = _cal_output_cell5()

    # 列的名称是2行，
    # print(result3)
    #
    # 0表示数值的第一行了
    # print(result4)
    head_rows = pd.DataFrame(
        [
            _build_cell_row(df, result3['cell3']),
            _build_cell_row(df, result4['cell4']),
            _build_cell_row(df, result5['cell5']),
        ],
        columns=df.columns
    )

    return pd.concat([head_rows, df], ignore_index=True)

# 字段进行对应 ，如timestamp改为time 类似
def cal_standardize_columns(df,params=None):
    output_columns = cal_ouput_columns()
    ds = pd.DataFrame(index=df.index, columns=output_columns)

    # Excel列 A 1 时间列
    ds[('Time', 'Time Step')] = df.index

    time_values = pd.to_datetime(df['Time'], format='%H:%M:%S')
    date_values = pd.to_datetime(df['Date'])

    # Excel列 B
    ds[('Time', 'Hour')] = time_values.dt.hour
    # Excel列 C
    ds[('Time', 'Day')] = date_values.dt.day
    # Excel列 D
    ds[('Time', 'Month')] = date_values.dt.month

    # Excel列 E PV列
    ds[('PV', 'AC MV Power Available')] = df['E_out']  #  E
    # Excel列 F 新增 - E-G-R, E 是负数则为0
    ds[('PV', 'AC MV Surplus')] = 0  # 光伏盈余 看 F

    # Excel列 G | 字段: pv_to_sub
    ds[('PV', 'AC MV Power to Infra')] = df['PV Power To Common Infrastructure Power [MW]'] # G

    # Excel列 H | 字段: pv_to_storage 新增 Excel列 H | 字段: pv_to_storage 充电时 H = K + R
    ds[('PV', 'AC MV Power to BESS')] = df['PV Power To BESS plant aux consumuer [MW]'] # H

    # Excel列 I | 字段: pv2grid  新增2 列的值不能<0,他不从电网取电
    ds[('PV', 'AC MV Power to HV')] = df['PV Power To Plant Substation BCP [MW]'] # I


    # Excel列 J J | 字段: pv2grid - 660   df['PV Power To Plant Substation BCP [MW]'] * params["eta_trafo"] * params['eta_cable'] # 来自光伏 AD
    ds[('PV', 'AC MV Power Dumped')] = df['PV Power To Plant Substation BCP [MW]'] * params["eta_trafo"] * params['eta_cable']  - params["pv_max_Discharge_cap"]  # 比660大的正的为结果值对把
    # Excel列 J
    ds[('PV', 'AC MV Power Dumped')] = ds[('PV', 'AC MV Power Dumped')].clip(lower=0) # 比660大的正的为结果值对把

    # Excel列 K BESS列 | 字段: Pin_ac  新增 -
    ds[('BESS', 'AC MV Power Charge')] = df['PV Power to BESS Plant BCP [MW]']  # 储能交流侧充电功率 K
    #  = pin_ac/( 8 * 6个效率 ) , eta_ch 是6个效率乘积  更改2 , 设置定值
    # ds[('BESS', 'AC MV Power Charge')] = df['PV Power to BESS Plant BCP [MW]']/(8*params['eta_ch'])  # 储能交流侧充电功率 K
    # ds[('BESS', 'AC MV Power Charge')] = 340

    # Excel列 L | 字段: Pdis_ac
    ds[('BESS', 'AC MV Power Discharge')] = df['DisCharge Power BCP [MW]']  # 储能交流侧放电功率 L
    # 更改2  = 高压侧目标放电功率 (MW)/(高压变压器效率 * 高压电缆效率 ) + 光伏 + infra 这里infra 是 PV Power To Common Infrastructure Power [MW]'
    ds[('BESS', 'AC MV Power Discharge')] = params["target_hv_power"]/(params['eta_trafo'] * params['eta_cable']) + ds[('PV', 'AC MV Power Available')] +  ds[('PV', 'AC MV Power to Infra')]


    # Excel列 M | 字段: Pin_dc
    # ds[('BESS', 'DC Power Charge')] = df['Charge Power [DC]']  # 储能直流侧充电功率 M
    # 更改2  = k * eta_ch 充电效率
    ds[('BESS', 'DC Power Charge')] = ds[('BESS', 'AC MV Power Charge')] * params['eta_ch']
    
    
    # Excel列 N | 字段: Pdis_dc
    ds[('BESS', 'DC Power Discharge')] = df['Discharge Power [DC]']  # 储能直流侧放电功率 N

    
    # Excel列 O | 字段: soc
    ds[('BESS', 'DC Energy')] = df['Energy [DC]']  # 储能直流侧能量  O
    # Excel列 P | 字段: soc / capacity
    ds[('BESS', 'SOC')] = df['SOC/%']  # 储能荷电状态 P

    # 映射 Mode，0：standby，1：charge，2：discharge
    mode_reverse_map = {'Standby': '0', 'Charging': '1', 'Discharge': '2'}

    # Excel列 Q
    ds[('BESS', 'Mode')] = df['Mode'].map(mode_reverse_map)  # 储能运行模式

    # Excel列 R 新增2 储能放电时 需要储能 提供的 - 未增  R | 字段: aux_storage，更改2 ，在原始计算中，区分充电8h，放电6的功率选择
    ds[('BESS', 'AC LV Aux Power')] = df['BESS Auxiliary Power [MW]']  # 储能辅助设备消耗功率

    # # Excel列 H | 字段: pv_to_storage 新增 Excel列 H | 字段: pv_to_storage 充电时 H = K + R
    # ds[('PV', 'AC MV Power to BESS')] = df['PV Power To BESS plant aux consumuer [MW]'] # H
    #
    #
    # # Excel列 H | 字段: pv_to_storage 新增 Excel列 H | 字段: pv_to_storage 充电时 H = K + R
    # ds[('BESS','_mask')] = ds[('BESS', 'AC MV Power Charge')] + ds[ ('BESS', 'AC LV Aux Power')]
    # mask = ds[ds[('BESS', 'Mode')] == '1'].index
    # ds[('PV', 'AC MV Power to BESS')] = [ds[('BESS','_mask')][i] if i in mask else ds[('PV', 'AC MV Power to BESS')][i]  for i in ds.index ]
    # del ds[('BESS','_mask')]

    # # Excel列 H  更改2 = k + R/(mvcab * mvtr 效率) AC MV Power to BESS
    ds[('PV', 'AC MV Power to BESS')] = ds[('BESS', 'AC MV Power Charge')] + ds[('BESS', 'AC LV Aux Power')]/(params['eta_trafo'] * params['eta_cable'])


    #  ======================== 更新  Excel列 F ---- F 列的数据 = E-G-R , E 是负数则为0 ====================================
    ds[('PV', 'AC MV Surplus')] = ds[('PV', 'AC MV Power Available')]-ds[('PV', 'AC MV Power to Infra')] - ds[('BESS', 'AC LV Aux Power')]
    ds[('PV', 'AC MV Surplus')] = ds[('PV', 'AC MV Surplus')].where(ds[('PV', 'AC MV Power Available')] > 0, 0)

    # Excel列 S 只有储能提供的时候，才会显示值 Q | 字段: bat_to_sub
    # ds[('BESS', 'AC MV Power to Infra')] = df['BESS plant Power To Common Infrastructure Power [MW]']  # 储能供给基础设施
    # 更改2 ，= 3
    ds[('BESS', 'AC MV Power to Infra')] = 3 # 储能供给基础设施


    # Excel列 T 加入 L列， # 没有更新，但是考虑的：放电功率 + 储能消耗辅助功率 ，不区分是谁提供的，只有放电的时候才会只有储能提供- 暂 | 字段: bat_to_storage
    # ds[('BESS', 'AC LV Power to BESS')] = df['BESS Power To BESS plant aux consumuer [MW]']  # 储能内部自耗或循环功率 T
    # ds[('BESS', 'AC LV Power to BESS')] = ds[('BESS', 'AC LV Power to BESS')] + ds[('BESS', 'AC MV Power Discharge')] # T
    # 更改2
    ds[('BESS', 'AC LV Power to BESS')] = ds[('BESS', 'DC Power Discharge')] * params["PCS Efficiency [%]:"] * params["LV Cable Efficiency [%]:"] * params["DC Cable Efficiency [%]:"] + ds[('BESS', 'AC LV Aux Power')]


    # Excel列 U | 字段: bat_to_pv
    # ds[('BESS', 'AC MV Power to PV')] = df['BESS Power To PV Plant [MW]']  # 储能向光伏馈电
    # 更改2 光伏不发电时的 abs(E),
    ds[('BESS', 'AC MV Power to PV')] = (-ds[('PV', 'AC MV Power Available')]).clip(lower=0) # 储能向光伏馈电


    # Excel列 V | 字段: storage2grid
    ds[('BESS', 'AC MV Power to HV')] = df['BESS Power To Plant Substation BCP [MW]']  # 储能向高压侧放电（如果没有可填 '-' 或 0）

    # Excel列 W HV列  W | 字段: pv2grid + storage2grid  新增2 列不应出现负值负值全部应为0
    ds[('HV', 'AC HV Power at MV of MV/HV Transformer')] = df['PV Power To Plant Substation BCP [MW]'] + df['BESS Power To Plant Substation BCP [MW]']  # 问题列 ： 高压侧交流中压功率 Mw不含高压变压器及高压电缆效率的功

    # Excel列 X 新增，需要有反向的数据， + -1*Z   ----- X | 字段: hv_power
    ds[('HV', 'AC HV Power at EDP')] = df['Exported Power [MW]']  # X 没有原始列可用，填 0 或 '-'

    # Excel列 Y 备用列 ----- Y | 字段: hv_power
    ds[('HV', 'Exported Power at EDP')] = df['Exported Power [MW]']

    # Excel列 Z 从电网输入功率 ----- Z | 字段: grid_power
    ds[('HV', 'Imported Power at EDP')] = df['Imported Power [MW]']

    # ======================== 更新 Excel列 X ----- X = Exported Power [MW] - Imported Power [MW]，hv_power - grid_power  相当于 = X - Z ===================
    ds[('HV', 'AC HV Power at EDP')] = ds[('HV', 'AC HV Power at EDP')] - ds[('HV', 'Imported Power at EDP')] # 新增，需要有反向的数据， + -1*Z

    # Excel列 AA 没有合并 | 字段: hv_power
    ds[('HVs', 'Exported Power at EDP.1')] = df['Exported Power [MW]']    # 备用列 AA

    # Excel列 AB | 字段: hv_power
    ds[('HVs', 'Exported Power > 1,000 MW')] = df['Exported Power [MW]']  # 超大功率输出 AB 无意义

    # 字段: storage2grid
    # ds[('HVs', 'Exported Power from BESS at EDP')] = df['BESS Power To Plant Substation BCP [MW]']  # 来自储能

    # # 新增   AA - AD
    # 字段: hv_power
    # ds[('HVs', 'Exported Power from BESS at EDP')] = df['Exported Power [MW]']  # 来自储能  AC
    #
    # # 新增 - AD   来自光伏 乘以 params["eta_trafo"] = 0.995    # 高压变压器效率    params["eta_cable"] = 0.99     # 高压电缆效率
    # 字段: pv2grid
    # ds[('HVs', 'Exported Power from PV at EDP')] = df['PV Power To Plant Substation BCP [MW]'] * params["eta_trafo"] * params['eta_cable'] # 来自光伏 AD
    # # AC = AA-AD
    # ds[('HVs', 'Exported Power from BESS at EDP')] = ds[('HVs', 'Exported Power at EDP.1')] - ds[('HVs', 'Exported Power from PV at EDP')]
    # 新增 - AD   来自光伏 乘以 params["eta_trafo"] = 0.995    # 高压变压器效率    params["eta_cable"] = 0.99     # 高压电缆效率

    # Excel列 AC  使用其他列 Exported Power [MW]， 如果 I>0 取值 Y 否则0 ， AC MV Power to HV 即 pv2grid>0 ,那么  取值 Y Exported Power [MW] 即 hv_power
    ds[('HVs', 'Exported Power from BESS at EDP')] = ds[('HV', 'Exported Power at EDP')].where(ds[('BESS', 'AC MV Power to HV')] > 0, 0)

    # Excel列 AD AD = AA-AC
    ds[('HVs', 'Exported Power from PV at EDP')] = ds[('HVs', 'Exported Power at EDP.1')]-ds[('HVs', 'Exported Power from BESS at EDP')]

    # 映射 Mode，0：standby，1：charge，2：discharge

    # 更新 excel 列 H, 充电的时候有值，其余为0
    ds[('PV', 'AC MV Power to BESS')] = ds[('PV', 'AC MV Power to BESS')].where(ds[('BESS', 'Mode')] == '1', 0)

    # Excel列 L 更改3 #  L = S + U +V + R * (params["MV Transformer Efficiency [%]:"] * params['MV Cable Efficiency [%]:'])
    ds[('BESS', 'AC MV Power Discharge')] = ds[('BESS', 'AC MV Power to Infra')] + ds[('BESS', 'AC MV Power to PV')] + ds[('BESS', 'AC MV Power to HV')]
    ds[('BESS', 'AC MV Power Discharge')] = ds[('BESS', 'AC MV Power Discharge')] + ds[('BESS', 'AC LV Aux Power')] * (params["MV Transformer Efficiency [%]:"] * params['MV Cable Efficiency [%]:'])


    # 更新 excel 列 L N T, 放电的时候有值，其余为0
    ds[('BESS', 'AC MV Power Discharge')] = ds[('BESS', 'AC MV Power Discharge')].where(ds[('BESS', 'Mode')] == '2', 0)


    # Excel列 N | 字段: Pdis_dc 更改3  Mode，0：standby，1：charge，2：discharge
    # [ L + S + U + R * (params["eta_trafo"] * params['eta_cable']) ]/ 放电效率
    # ds[('BESS', 'DC Power Discharge')] = ds[('BESS', 'AC MV Power Discharge')] +  ds[('BESS', 'AC MV Power to Infra')] + ds[('BESS', 'AC MV Power to PV')] + ds[('BESS', 'AC LV Aux Power')] * (params["MV Transformer Efficiency [%]:"] * params['MV Cable Efficiency [%]:'])
    # ds[('BESS', 'DC Power Discharge')] = ds[('BESS', 'DC Power Discharge')]/params['eta_dis']
    #  更新  N  =  L / 放电效率
    ds[('BESS', 'DC Power Discharge')] = ds[('BESS', 'AC MV Power Discharge')]/params['eta_dis']
    ds[('BESS', 'DC Power Discharge')] = ds[('BESS', 'DC Power Discharge')].where(ds[('BESS', 'Mode')] == '2', 0)

    # 更新 excel列 I  新增3 = F-G - k - R/(params["eta_trafo"] * params['eta_cable'])
    ds[('PV', 'AC MV Power to HV')] = ds[('PV', 'AC MV Surplus')] -  ds[('PV', 'AC MV Power to Infra')] - ds[('BESS', 'AC MV Power Charge')] - ds[('BESS', 'AC LV Aux Power')]/(params["eta_trafo"] * params['eta_cable'])
    ds[('PV', 'AC MV Power to HV')] = ds[('PV', 'AC MV Power to HV')].where(ds[('PV', 'AC MV Power Available')] > 0, 0)


    # Excel列 T  更改3 放电状态下：T=R ，其他未0
    ds[('BESS', 'AC LV Power to BESS')] = ds[('BESS', 'AC LV Aux Power')]
    ds[('BESS', 'AC LV Power to BESS')] = ds[('BESS', 'AC LV Power to BESS')].where(ds[('BESS', 'Mode')] == '2', 0)

    # Excel列 N  更改3  = L / eta_dis 放电效率
    ds[('BESS', 'DC Power Discharge')] = ds[('BESS', 'AC MV Power Discharge')] / params['eta_dis']


    return ds[output_columns]


# 数据拼接
    # print(f"✅ 所有文件处理完成，保存到 {output_file}")

# 格式设置
def format_all_sheets_with_method(input_file):
    # -------------------置（保持不变）-------------------
    fill_config = [(1, 5, 1, 1, 'F2F2F2'), (1, 5, 2, 2, 'F2F2F2'), (1, 5, 3, 3, 'F2F2F2'), (1, 5, 4, 4, 'F2F2F2'),
                   (1, 5, 5, 5, 'E2F0D9'), (1, 5, 6, 6, 'E2F0D9'), (1, 5, 7, 7, 'E2F0D9'), (1, 5, 8, 8, 'E2F0D9'),
                   (1, 5, 9, 9, 'E2F0D9'), (1, 5, 10, 10, 'E2F0D9'), (1, 5, 11, 11, 'D9E1F2'), (1, 5, 12, 12, 'D9E1F2'),
                   (1, 5, 13, 13, 'D9E1F2'), (1, 5, 14, 14, 'D9E1F2'), (1, 5, 15, 15, 'D9E1F2'),
                   (1, 5, 16, 16, 'D9E1F2'),
                   (1, 5, 17, 17, 'D9E1F2'), (1, 5, 18, 18, 'D9E1F2'), (1, 5, 19, 19, 'D9E1F2'),
                   (1, 5, 20, 20, 'D9E1F2'),
                   (1, 5, 21, 21, 'D9E1F2'), (1, 5, 22, 22, 'D9E1F2'), (1, 5, 23, 23, 'FFF2CC'),
                   (1, 5, 24, 24, 'FFF2CC'),
                   (1, 5, 25, 25, 'FFF2CC'), (1, 5, 26, 26, 'FFF2CC')]

    font_config = [(27, '#0070C0', 9, True), (28, '#0070C0', 9, True), (29, '#0070C0', 9, True),
                   (30, '#0070C0', 9, True)]
    underline_config = [(5, 1, 30)]
    right_border_config = [(4, 1, 2000)]
    merge_config = [(1, 1, 1, 4, None), (1, 5, 1, 10, 'PV'), (1, 11, 1, 22, 'BESS'), (1, 23, 1, 26, 'HV'),
                    (1, 27, 1, 30, None)]
    left_align_rows = [(2, 5, 1, 30)]
    col_width = {1: 8, 2: 8, 3: 8, 4: 8, 5: 8, 6: 8, 7: 12, 8: 8, 9: 8, 10: 8, 11: 12, 12: 12, 13: 9, 14: 9, 15: 9,
                 16: 6, 17: 6, 18: 12, 19: 12, 20: 12, 21: 9, 22: 9, 23: 8, 24: 8, 25: 9, 26: 9, 27: 9, 28: 9, 29: 9,
                 30: 9}

    output_file = input_file.replace('.xlsx', '_format.xlsx')
    start = time.time()
    format_excel_all_sheets(
        input_file,
        output_file,
        fill_config=fill_config,
        font_config=font_config,
        underline_config=underline_config,
        right_border_config=right_border_config,
        merge_config=merge_config,
        left_align_rows=left_align_rows,
        col_width=col_width
    )
    end = time.time()
    # print(f"执行时间: {end - start:.3f} 秒")
    print(f"✅ 所有 sheet 格式处理完成：{output_file}")


    # 旧逻辑保留，不删除。

def _output_format_config():
    return {
        "fill_config": [(1, 5, 1, 1, 'F2F2F2'), (1, 5, 2, 2, 'F2F2F2'), (1, 5, 3, 3, 'F2F2F2'), (1, 5, 4, 4, 'F2F2F2'),
                        (1, 5, 5, 5, 'E2F0D9'), (1, 5, 6, 6, 'E2F0D9'), (1, 5, 7, 7, 'E2F0D9'), (1, 5, 8, 8, 'E2F0D9'),
                        (1, 5, 9, 9, 'E2F0D9'), (1, 5, 10, 10, 'E2F0D9'), (1, 5, 11, 11, 'D9E1F2'), (1, 5, 12, 12, 'D9E1F2'),
                        (1, 5, 13, 13, 'D9E1F2'), (1, 5, 14, 14, 'D9E1F2'), (1, 5, 15, 15, 'D9E1F2'),
                        (1, 5, 16, 16, 'D9E1F2'), (1, 5, 17, 17, 'D9E1F2'), (1, 5, 18, 18, 'D9E1F2'),
                        (1, 5, 19, 19, 'D9E1F2'), (1, 5, 20, 20, 'D9E1F2'), (1, 5, 21, 21, 'D9E1F2'),
                        (1, 5, 22, 22, 'D9E1F2'), (1, 5, 23, 23, 'FFF2CC'), (1, 5, 24, 24, 'FFF2CC'),
                        (1, 5, 25, 25, 'FFF2CC'), (1, 5, 26, 26, 'FFF2CC')],
        "font_config": [(27, '#0070C0', 9, True), (28, '#0070C0', 9, True), (29, '#0070C0', 9, True), (30, '#0070C0', 9, True)],
        "underline_config": [(5, 1, 30)],
        "right_border_config": [(4, 1, 2000)],
        "merge_config": [(1, 1, 1, 4, None), (1, 5, 1, 10, 'PV'), (1, 11, 1, 22, 'BESS'), (1, 23, 1, 26, 'HV'),
                         (1, 27, 1, 30, None)],
        "left_align_rows": [(2, 5, 1, 30)],
        "col_width": {1: 8, 2: 8, 3: 8, 4: 8, 5: 8, 6: 8, 7: 12, 8: 8, 9: 8, 10: 8, 11: 12, 12: 12, 13: 9, 14: 9, 15: 9,
                      16: 6, 17: 6, 18: 12, 19: 12, 20: 12, 21: 9, 22: 9, 23: 8, 24: 8, 25: 9, 26: 9, 27: 9, 28: 9,
                      29: 9, 30: 9},
    }

def _xlsx_value(value):
    if pd.isna(value):
        return ""
    return value

def _apply_output_sheet_format(workbook, ws, sheet_values, config):
    nrows, ncols = sheet_values.shape

    col_width = config["col_width"]
    for col_idx, width in col_width.items():
        ws.set_column(col_idx - 1, col_idx - 1, width)

    format_cache = {}
    def get_format(props):
        key = tuple(sorted(props.items()))
        if key not in format_cache:
            format_cache[key] = workbook.add_format(props)
        return format_cache[key]

    cell_styles = {}
    for r1, r2, c1, c2, color in config["fill_config"]:
        for row in range(r1 - 1, min(r2, nrows)):
            for col in range(c1 - 1, min(c2, ncols)):
                cell_styles.setdefault((row, col), {}).update({'bg_color': color})

    for col_idx, font_color, font_size, bold in config["font_config"]:
        col = col_idx - 1
        if col < ncols:
            for row in range(nrows):
                cell_styles.setdefault((row, col), {}).update({
                    'font_color': font_color, 'font_size': font_size, 'bold': bold
                })

    for row_idx, col_start, col_end in config["underline_config"]:
        row = row_idx - 1
        if row < nrows:
            for col in range(col_start - 1, min(col_end, ncols)):
                cell_styles.setdefault((row, col), {}).update({'bottom': 1})

    for col_idx, row_start, row_end in config["right_border_config"]:
        col = col_idx - 1
        if col < ncols:
            for row in range(row_start - 1, min(row_end, nrows)):
                cell_styles.setdefault((row, col), {}).update({'right': 1})

    for r1, r2, c1, c2 in config["left_align_rows"]:
        for row in range(r1 - 1, min(r2, nrows)):
            for col in range(c1 - 1, min(c2, ncols)):
                cell_styles.setdefault((row, col), {}).update({
                    'align': 'left', 'valign': 'vcenter', 'text_wrap': True
                })

    for (row, col), props in cell_styles.items():
        ws.write(row, col, _xlsx_value(sheet_values[row, col]), get_format(props))

    for r1, c1, r2, c2, value in config["merge_config"]:
        row_start, col_start = r1 - 1, c1 - 1
        row_end, col_end = r2 - 1, c2 - 1
        if row_start >= nrows or col_start >= ncols:
            continue
        merge_props = {'align': 'center', 'valign': 'vcenter'}
        merge_props.update(cell_styles.get((row_start, col_start), {}))
        ws.merge_range(row_start, col_start, row_end, col_end, value, get_format(merge_props))

def _write_output_sheet_xlsxwriter(workbook, sheet_name, df_year, config):
    ws = workbook.add_worksheet(sheet_name)
    header_rows = [
        [str(lvl0) for lvl0, _ in df_year.columns],
        [str(lvl1) for _, lvl1 in df_year.columns],
    ]
    data_values = df_year.to_numpy(dtype=object, copy=False)
    sheet_values = np.vstack([np.array(header_rows, dtype=object), data_values])

    for row_idx, row in enumerate(sheet_values):
        ws.write_row(row_idx, 0, [_xlsx_value(value) for value in row])

    _apply_output_sheet_format(workbook, ws, sheet_values, config)

def _save_formatted_year_raw_dfs_to_workbook(all_year_raw_dfs, output_file, sheet_prefix):
    if os.path.exists(output_file):
        os.remove(output_file)

    config = _output_format_config()
    workbook = xlsxwriter.Workbook(output_file)
    try:
        for item in all_year_raw_dfs:
            year_index = item["year_index"]
            df_raw = item["df"]
            params = item["params"]

            df_year = _deal_save_data_df_pv_sheet1(df_raw, params)
            df_year = cal_standardize_columns(df_year, params=params)
            df_year = cal_output_head(df_year)

            sheet_name = f"{sheet_prefix}{str(year_index).zfill(2)}"
            _write_output_sheet_xlsxwriter(workbook, sheet_name, df_year, config)
    finally:
        workbook.close()

    print(f"所有 sheet 已直接保存为格式化文件：{output_file}")


def save_all_year_raw_dfs_to_single_workbook(
        all_year_raw_dfs=None,
        output_file=None,
        sheet_prefix="Y",
        format_after_save=True,
):
    """
    根据 run_phase 收集的原始年度 df 列表，一次性生成最终大表。
    不再重复从磁盘读取每年的 xlsx。
    """
    if not all_year_raw_dfs:
        raise ValueError("all_year_raw_dfs 不能为空")
    if output_file is None:
        raise ValueError("output_file 不能为空")

    if format_after_save:
        formatted_output_file = output_file.replace('.xlsx', '_format.xlsx')
        _save_formatted_year_raw_dfs_to_workbook(all_year_raw_dfs, formatted_output_file, sheet_prefix)
        return

    wb = Workbook(write_only=True)
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    if os.path.exists(output_file):
        os.remove(output_file)

    for item in all_year_raw_dfs:
        year_index = item["year_index"]
        df_raw = item["df"].copy()
        params = item["params"]

        df_year = _deal_save_data_df_pv_sheet1(df_raw,params)
        df_year = cal_standardize_columns(df_year, params=params)
        df_year = cal_output_head(df_year)

        sheet_name = f"{sheet_prefix}{str(year_index).zfill(2)}"
        ws = wb.create_sheet(sheet_name)

        ws.append([str(lvl0) for lvl0, _ in df_year.columns])
        ws.append([str(lvl1) for _, lvl1 in df_year.columns])

        for row in df_year.itertuples(index=False, name=None):
            ws.append(row)

    wb.save(output_file)
    wb.close()

# format_all_sheets_with_method(R"E:\bess_shows\PVsyst数据\结果数据\UAE_DEWA7_Phase_A_PM_Output v1.0.xlsx")
# format_all_sheets_with_method(R"E:\bess_shows\PVsyst数据\结果数据\UAE_DEWA7_Phase_B_PM_Output v1.0.xlsx")
# format_all_sheets_with_method(R"E:\bess_shows\PVsyst数据\结果数据\UAE_DEWA7_Phase_C_PM_Output v1.0.xlsx")
